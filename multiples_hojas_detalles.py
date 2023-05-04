
import pandas as pd
import mysql.connector
import openpyxl


def connectionBD():
    try:
        connection = mysql.connector.connect(
            host="localhost",
            user="root",
            passwd="",
            database="demo",
            raise_on_warnings=True
        )
        if connection.is_connected():
            return connection

    except mysql.connector.Error as error:
        print(f"No se pudo conectar: {error}")


def listaPersonas():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = "SELECT * FROM personas"
                mycursor.execute(querySQL)
                resumen_personas = mycursor.fetchall()
                return resumen_personas
    except Exception as e:
        print(f"Ocurrió un error leyendo las personas: {e}")
        return {}


def listaUsuarios():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = "SELECT * FROM registros"
                mycursor.execute(querySQL)
                resumen_usuarios = mycursor.fetchall()
                return resumen_usuarios
    except Exception as e:
        print(f"Ocurrió un error leyendo los usuarios: {e}")
        return {}


def hojaUsuarios(lista_usuarios, wb):
    # Crea una hoja de Excel para los usuarios
    hoja = wb.create_sheet(title="Usuarios")
    # Crea la fila del encabezado con los títulos
    hoja.append(('id', 'nombre', 'profesion', 'status'))
    for usuario in lista_usuarios:
        # Agrega una tupla con los valores del usuario
        hoja.append((usuario['id'], usuario['nombre'],
                    usuario['profesion'], usuario['status']))


def hojaPersonas(lista_personas, wb):
    # Crea una hoja de Excel para las personas
    hoja = wb.create_sheet(title="Personas")
    # Crea la fila del encabezado con los títulos
    hoja.append(('id', 'Nombre', 'Sexo', 'email'))
    for persona in lista_personas:
        # Agrega una tupla con los valores de la persona
        hoja.append((persona['id_persona'], persona['nombre'],
                    persona['sexo'], persona['email']))


def resumenData():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = "SELECT * FROM personas"
                mycursor.execute(querySQL)
                resumen_personas = mycursor.fetchall()

                # Crea un archivo Excel
                wb = openpyxl.Workbook()

                # Elimina la hoja en blanco llamada "Sheet"
                wb.remove(wb.active)

                # Crea una hoja de Excel para las personas
                hojaPersonas(resumen_personas, wb)

                # Crea una hoja de Excel para los usuarios
                resumen_usuarios = listaUsuarios()
                hojaUsuarios(resumen_usuarios, wb)

                # Guarda el archivo Excel
                wb.save('resumenData.xlsx')

    except Exception as e:
        print(f"Ocurrió un error leyendo las consignaciones: {e}")
        return {}


if __name__ == '__main__':
    resumenData()
