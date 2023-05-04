import pandas as pd
import mysql.connector
import openpyxl


# Funcion para establecer conexión con BD MySQL y Python
def connectionBD():
    try:
        connection = mysql.connector.connect(
            host="localhost",
            user="root",
            passwd="",
            database="demo",
            raise_on_warnings=True
        )
        if connection.is_connected():
            return connection

    except mysql.connector.Error as error:
        print(f"No se pudo conectar: {error}")


# Función que genera la 2 hoja del Excel para la tabla de Usuarios
def listaPersonasBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = "SELECT * FROM personas"
                mycursor.execute(querySQL)
                lista_personas = mycursor.fetchall()
                return lista_personas
    except Exception as e:
        print(f"Ocurrió un error leyendo las personas: {e}")
        return {}


# Función que genera la 2 hoja del Excel para la tabla de Registros
def listaRegistrosUsuarios():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = "SELECT * FROM registros"
                mycursor.execute(querySQL)
                return mycursor.fetchall()
    except Exception as e:
        print(f"Ocurrió un error {e}")
        return {}


# Función para crear la Hoja de Registros Usuarios
def hojaRegistrosUsuarios(lista_registros_usuarios, wb):
    # Crea una hoja de Excel para los usuarios
    hoja = wb.create_sheet(title="Registros Usuarios")
    # Crea la fila del encabezado con los títulos
    hoja.append(('Id', 'Nombre', 'Profesion', 'Estatus'))
    for registro in lista_registros_usuarios:
        # Agrega una tupla con los valores del usuario
        hoja.append((registro['id'], registro['nombre'],
                    registro['profesion'], registro['status']))


# Función para generar la hoja de Personas
def hojaPersonas(lista_personas, wb):
    # Crea una hoja de Excel para las personas
    hoja = wb.create_sheet(title="Personas")
    # Crea la fila del encabezado con los títulos
    hoja.append(('Id', 'Nombre', 'Sexo', 'Email'))
    for persona in lista_personas:
        # Agrega una tupla con los valores de la persona
        hoja.append((persona['id_persona'], persona['nombre'],
                    persona['sexo'], persona['email']))


# Funcion para consolidad ambas consultar y generar el Excel
def resumenData():
    try:
        # Crea un archivo Excel
        wb = openpyxl.Workbook()

        # Elimina la hoja en blanco llamada "Sheet"
        wb.remove(wb.active)

        # Crea una hoja de Excel para las personas
        hojaPersonas(listaPersonasBD(), wb)

        # Crea una hoja de Excel para los usuarios
        hojaRegistrosUsuarios(listaRegistrosUsuarios(), wb)

        # Guarda el archivo Excel
        wb.save('resumenData_3.xlsx')

    except Exception as e:
        print(f"Ocurrió un error leyendo las consignaciones: {e}")
        return {}


if __name__ == '__main__':
    resumenData()
